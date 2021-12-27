from openpyxl import *
from utils.PullData import PullData
from utils.PullDoi import PullDoi
from utils.Creating_Edge import *

if __name__ == '__main__':
    wb = load_workbook('digCon.xlsx')
    ws = wb.active
    doirecord = []  #change variable name here to avoid conflicts between the doi in line20
    noDOI = []

    for i in range(5, 6):
        d = ws.cell(row = i, column = 12).value
        tit = ws.cell(row = i, column = 4).value
        if d:
            doirecord.append(d)
        else:
            noDOI.append(tit)

    g = Creating_Graph(doirecord)
    g.Create_Graph()

    # for dx in doirecord:
    #
    #     item = PullData(dx)
    #     doi, typ, vol, iss, pgs, aut, tit, pub, cit, jnl, ref, ymd, issn, isbn, loc, url, edt, anm = item.pullData()

    # for dx in noDOI:
    #     pdoi = PullDoi(dx)
    #     item = PullData(pdoi)
    #     doi, typ, vol, iss, pgs, aut, tit, pub, cit, jnl, ref, ymd, issn, isbn, loc, url, edt, anm = item.pullData()
    #     doirecord.append(doi)
    print("  ")